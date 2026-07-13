"""Б4: опись личных дел уволенных из отчёта 1С:ЗиК «Принято уволено».

Правила брифа УРП:
- в опись попадают только уволенные БЕЗ повторного приёма за период
  (у сотрудника нет события «Прием» после увольнения);
- дата увольнения = «Дата записи о гражданстве» − 1 календарный день;
- по умолчанию — категории административного блока (АУП/АХП/УВП, опись так
  и называется), остальных можно включить флагом all_categories;
- результат — xlsx по образцу: шапка ТИУ, УТВЕРЖДАЮ, таблица
  (№, ФИО, Должность (ставка), Подразделение, Дата увольнения), подписи.

Всё детерминированно, без LLM.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta
from pathlib import Path

from sqlalchemy.orm import Session

from config import settings
from data.my_documents import MyDocuments
from data.users import User
from utils.logger import logger

# Триггер чат-команды: «сделай опись личных дел уволенных»
INVENTORY_REQUEST_RE = re.compile(
    r"опис\w*\s+(?:личн\w+\s+дел|увол)|уволенн\w+\s+в\s+архив",
    re.IGNORECASE,
)

# Категории «административно уволенных» из названия описи
_ADMIN_CATEGORIES = {"АУП", "АУПН", "АХП", "АХПН", "УВП"}
_DT_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")


def _cell_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = _DT_RE.search(str(v or ""))
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def analyze_dismissed_xls(path: str | Path, all_categories: bool = False) -> dict:
    """Разбирает отчёт «Принято уволено» и строит список для описи."""
    from openpyxl import load_workbook

    p = Path(path)
    if p.suffix.lower() == ".xls":
        from services.parsers.office_convert import convert_to_modern

        p = convert_to_modern(p)
    ws = load_workbook(str(p), data_only=True).active

    # Шапка: строка со столбцами «Сотрудник» и «Вид события»
    header_row = None
    col: dict[str, int] = {}
    for ri, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        cells = [str(v).strip().lower() if v is not None else "" for v in row]
        if any("сотрудник" in c for c in cells) and any("вид события" in c for c in cells):
            header_row = ri
            for ci, c in enumerate(cells):
                if "дата записи" in c:
                    col["record_date"] = ci
                elif c == "сотрудник":
                    col["fio"] = ci
                elif "иерарх" in c:
                    col["unit"] = ci
                elif c == "должность":
                    col["position"] = ci
                elif "категория" in c:
                    col["category"] = ci
                elif "количество ставок" in c:
                    col["rate"] = ci
                elif "вид события" in c:
                    col["event"] = ci
            break
    if header_row is None or "fio" not in col or "event" not in col:
        raise ValueError("Файл не похож на отчёт «Принято уволено» из 1С:ЗиК")

    def _get(row, key):
        i = col.get(key)
        v = row[i] if i is not None and i < len(row) else None
        return v if v is not None else ""

    records: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        fio = str(_get(row, "fio")).strip()
        event = str(_get(row, "event")).strip().lower()
        if not event:
            continue
        rec_date = _cell_date(_get(row, "record_date"))
        records.append({
            "fio": fio,
            "unit": str(_get(row, "unit")).strip(),
            "position": str(_get(row, "position")).strip(),
            "category": str(_get(row, "category")).strip().upper(),
            "rate": str(_get(row, "rate")).strip(),
            "event": "hire" if "прием" in event or "приём" in event else
                     ("fire" if "увольнение" in event else event),
            "record_date": rec_date,
        })
    if not records:
        raise ValueError("В отчёте не нашлось строк с событиями «Прием»/«Увольнение»")

    # Повторный приём: у сотрудника есть «Прием» с датой ПОЗЖЕ увольнения
    # (или без даты — консервативно тоже считаем повторным приёмом).
    by_fio: dict[str, list[dict]] = {}
    for r in records:
        if r["fio"]:
            by_fio.setdefault(r["fio"].lower(), []).append(r)

    items: list[dict] = []
    skipped_rehired = 0
    for r in records:
        if r["event"] != "fire":
            continue
        rehired = False
        for other in by_fio.get(r["fio"].lower(), []):
            if other["event"] != "hire":
                continue
            if other["record_date"] is None or r["record_date"] is None or \
                    other["record_date"] >= r["record_date"]:
                rehired = True
                break
        if rehired:
            skipped_rehired += 1
            continue
        if not all_categories and r["category"] and r["category"] not in _ADMIN_CATEGORIES:
            continue
        dismiss = r["record_date"] - timedelta(days=1) if r["record_date"] else None
        pos = r["position"]
        if r["rate"]:
            pos = f"{pos} ({r['rate']} ст.)" if pos else f"({r['rate']} ст.)"
        items.append({
            "fio": r["fio"] or "—",
            "position": pos,
            "unit": r["unit"],
            "dismissed_at": dismiss.strftime("%d.%m.%Y") if dismiss else "",
            "_sort": (r["fio"].lower(), dismiss or date.min),
        })

    items.sort(key=lambda x: x["_sort"])
    for i, it in enumerate(items, 1):
        it.pop("_sort")
        it["n"] = i

    years = [d["record_date"].year for d in records if d["record_date"]]
    year = max(set(years), key=years.count) if years else date.today().year
    return {
        "year": year,
        "total_records": len(records),
        "fired_total": len([r for r in records if r["event"] == "fire"]),
        "skipped_rehired": skipped_rehired,
        "items": items,
    }


_UNITS = ["", "одно", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
_TEENS = ["десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать",
          "пятнадцать", "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать"]
_TENS = ["", "", "двадцать", "тридцать", "сорок", "пятьдесят", "шестьдесят",
         "семьдесят", "восемьдесят", "девяносто"]
_HUNDREDS = ["", "сто", "двести", "триста", "четыреста", "пятьсот", "шестьсот",
             "семьсот", "восемьсот", "девятьсот"]


def _num_words(n: int) -> str:
    """0–999 прописью (для «157 (сто пятьдесят семь) личных дел»)."""
    if n == 0:
        return "ноль"
    parts = [_HUNDREDS[n // 100]]
    rem = n % 100
    if 10 <= rem <= 19:
        parts.append(_TEENS[rem - 10])
    else:
        parts.append(_TENS[rem // 10])
        parts.append(_UNITS[rem % 10])
    return " ".join(p for p in parts if p)


def build_inventory_xlsx(result: dict) -> bytes:
    """xlsx-опись по образцу УРП (шапка, УТВЕРЖДАЮ, таблица, подписи)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Опись"
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _merged(row: int, text: str, bold=False, size=11, align=center):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row, 1, text)
        c.font = Font(bold=bold, size=size)
        c.alignment = align

    _merged(1, "МИНИСТЕРСТВО НАУКИ И ВЫСШЕГО ОБРАЗОВАНИЯ РОССИЙСКОЙ ФЕДЕРАЦИИ", size=10)
    _merged(2, "ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ ВЫСШЕГО ОБРАЗОВАНИЯ", size=10)
    _merged(3, "«ТЮМЕНСКИЙ ИНДУСТРИАЛЬНЫЙ УНИВЕРСИТЕТ»", bold=True, size=11)
    _merged(4, "УПРАВЛЕНИЕ ПО РАБОТЕ С ПЕРСОНАЛОМ", size=10)
    ws.cell(6, 4, "УТВЕРЖДАЮ").font = Font(bold=True)
    ws.cell(7, 4, "Начальник УРП")
    ws.cell(8, 4, "____________ Н. Г. Дударева")
    ws.cell(9, 4, "«____» __________ 20___ г.")
    _merged(11,
            "ОПИСЬ\nличных дел административно-управленческого, административно-хозяйственного "
            f"и учебно-вспомогательного персонала ТИУ, уволенных в {result['year']} году",
            bold=True)
    ws.row_dimensions[11].height = 48

    hdr_row = 13
    headers = ["№ п/п", "Ф.И.О.", "Должность", "Структурное подразделение", "Дата увольнения"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = Font(bold=True)
        c.alignment = center
        c.border = border
    for it in result["items"]:
        r = hdr_row + it["n"]
        for ci, v in enumerate([it["n"], it["fio"], it["position"], it["unit"], it["dismissed_at"]], 1):
            c = ws.cell(r, ci, v)
            c.border = border
            c.alignment = left if ci in (2, 3, 4) else center

    n = len(result["items"])
    foot = hdr_row + n + 2
    ws.cell(foot, 1, f"Передала {n} ({_num_words(n)}) личных дел")
    ws.cell(foot + 1, 1, "Начальник УРП")
    ws.cell(foot + 1, 4, "Н. Г. Дударева")
    ws.cell(foot + 3, 1, f"Приняла {n} ({_num_words(n)}) личных дел")
    ws.cell(foot + 4, 1, "Руководитель архивной службы общего отдела")
    ws.cell(foot + 4, 4, "О. И. Вологодская")

    for letter, width in (("A", 7), ("B", 32), ("C", 34), ("D", 40), ("E", 15)):
        ws.column_dimensions[letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_inventory(
    db: Session, user: User, xls_path: str | Path, all_categories: bool = False
) -> tuple[MyDocuments, dict]:
    result = analyze_dismissed_xls(xls_path, all_categories=all_categories)
    settings.docs_generated.mkdir(parents=True, exist_ok=True)
    out = settings.docs_generated / f"inventory_{result['year']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out.write_bytes(build_inventory_xlsx(result))
    rec = MyDocuments(
        user_id=user.id,
        title=f"Опись личных дел уволенных в {result['year']} году ({len(result['items'])} чел.)",
        template_key="dismissed_inventory",
        file_path=str(out),
        progress=100,
        status="ready",
        fields={"year": result["year"], "count": len(result["items"])},
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    logger.info("[INVENTORY] опись создана: doc={} ({} чел.)", rec.id, len(result["items"]))
    return rec, result
