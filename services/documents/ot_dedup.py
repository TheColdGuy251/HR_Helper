"""Б7: поиск однотипных инструкций по охране труда.

Отдел ОТ сокращает 448 действующих инструкций: нужно найти пары документов,
у которых «более 80% текста совпадает». Двухступенчатый алгоритм без LLM:

1) дешёвый кандидат-отбор: шинглы (5-словные n-граммы, хеши) + инвертированный
   индекс → Dice-коэффициент по множествам шинглов для всех пар сразу;
2) точная похожесть для кандидатов (Dice ≥ 0.5): difflib.SequenceMatcher
   по СЛОВАМ (быстро и близко к человеческому «процент совпадения текста»).

Пары ≥ порога объединяются в группы (связные компоненты) — так видно целые
семейства однотипных инструкций, а не только парные совпадения.
"""

from __future__ import annotations

import io
import re
import zlib
from collections import defaultdict
from difflib import SequenceMatcher

from utils.logger import logger

# Триггер чат-команды: «найди дубликаты/однотипные инструкции»
OT_DEDUP_REQUEST_RE = re.compile(
    r"(?:дубл|однотип|совпаден)[\w-]*[^.]{0,60}инструкц|инструкц[\w-]*[^.]{0,60}(?:дубл|однотип)",
    re.IGNORECASE,
)

_WORD_RE = re.compile(r"[а-яёa-z0-9]+", re.IGNORECASE)
_SHINGLE = 5
# Кандидаты на точное сравнение (быстрый Dice по шинглам)
_CANDIDATE_DICE = 0.5
# Порог «однотипные» из брифа отдела ОТ
DUPLICATE_THRESHOLD = 0.80
# Пары ниже порога, но выше этого — «пограничные», тоже попадают в отчёт
REPORT_THRESHOLD = 0.60


def _words(text: str) -> list[str]:
    return _WORD_RE.findall((text or "").lower())


def _shingles(words: list[str]) -> set[int]:
    if len(words) < _SHINGLE:
        return {zlib.crc32(" ".join(words).encode("utf-8"))} if words else set()
    return {
        zlib.crc32(" ".join(words[i:i + _SHINGLE]).encode("utf-8"))
        for i in range(len(words) - _SHINGLE + 1)
    }


def compare_documents(docs: list[tuple[str, str]]) -> dict:
    """docs: [(имя файла, текст)] → {pairs, groups, files}.

    pairs: [{a, b, percent}] по убыванию похожести (percent — пословный
    SequenceMatcher, ≥ REPORT_THRESHOLD); groups: связные компоненты пар
    ≥ DUPLICATE_THRESHOLD с диапазоном похожести внутри группы.
    """
    names = [n for n, _ in docs]
    word_lists = [_words(t) for _, t in docs]
    shingle_sets = [_shingles(w) for w in word_lists]

    # Инвертированный индекс шинглов → счётчик общих шинглов по парам
    by_shingle: dict[int, list[int]] = defaultdict(list)
    for i, ss in enumerate(shingle_sets):
        for h in ss:
            by_shingle[h].append(i)
    common: dict[tuple[int, int], int] = defaultdict(int)
    for ids in by_shingle.values():
        if len(ids) < 2 or len(ids) > 50:  # шингл-«шум», общий для всех, не информативен
            continue
        for x in range(len(ids)):
            for y in range(x + 1, len(ids)):
                common[(ids[x], ids[y])] += 1

    candidates: list[tuple[int, int]] = []
    for (i, j), c in common.items():
        denom = len(shingle_sets[i]) + len(shingle_sets[j])
        if denom and (2 * c / denom) >= _CANDIDATE_DICE:
            candidates.append((i, j))
    logger.info("[OT-DEDUP] файлов: {}, пар-кандидатов: {}", len(docs), len(candidates))

    pairs: list[dict] = []
    for i, j in candidates:
        ratio = SequenceMatcher(a=word_lists[i], b=word_lists[j], autojunk=False).ratio()
        if ratio >= REPORT_THRESHOLD:
            pairs.append({
                "a": names[i], "b": names[j],
                "percent": round(ratio * 100, 1),
                "_i": i, "_j": j,
            })
    pairs.sort(key=lambda p: -p["percent"])

    # Группы однотипных (связные компоненты по парам ≥ 80%)
    parent = list(range(len(docs)))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for p in pairs:
        if p["percent"] >= DUPLICATE_THRESHOLD * 100:
            a, b = find(p["_i"]), find(p["_j"])
            if a != b:
                parent[a] = b

    members: dict[int, list[int]] = defaultdict(list)
    for i in range(len(docs)):
        members[find(i)].append(i)
    groups: list[dict] = []
    for ids in members.values():
        if len(ids) < 2:
            continue
        in_group = [
            p["percent"] for p in pairs
            if p["percent"] >= DUPLICATE_THRESHOLD * 100
            and find(p["_i"]) == find(ids[0])
        ]
        groups.append({
            "files": sorted(names[i] for i in ids),
            "size": len(ids),
            "min_percent": min(in_group) if in_group else 0,
            "max_percent": max(in_group) if in_group else 0,
        })
    groups.sort(key=lambda g: -g["size"])

    for p in pairs:
        p.pop("_i"), p.pop("_j")
    return {
        "files": len(docs),
        "pairs": pairs,
        "duplicates": len([p for p in pairs if p["percent"] >= DUPLICATE_THRESHOLD * 100]),
        "groups": groups,
    }


def build_dedup_xlsx(result: dict) -> bytes:
    """xlsx-отчёт: лист «Пары» (все ≥60%) + лист «Группы однотипных» (≥80%)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Пары"
    ws.append(["Инструкция 1", "Инструкция 2", "Совпадение, %"])
    for c in ws[1]:
        c.font = Font(bold=True)
    red = PatternFill("solid", start_color="FDE8E8")
    for p in result["pairs"]:
        ws.append([p["a"], p["b"], p["percent"]])
        if p["percent"] >= DUPLICATE_THRESHOLD * 100:
            for c in ws[ws.max_row]:
                c.fill = red
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 16

    ws2 = wb.create_sheet("Группы однотипных")
    ws2.append(["Группа", "Файлов", "Совпадение, %", "Файлы (кандидаты на объединение)"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for gi, g in enumerate(result["groups"], 1):
        rng = (
            f"{g['min_percent']:.0f}–{g['max_percent']:.0f}"
            if g["min_percent"] != g["max_percent"] else f"{g['max_percent']:.0f}"
        )
        ws2.append([gi, g["size"], rng, "\n".join(g["files"])])
        ws2.cell(ws2.max_row, 4).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["D"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run_dedup_zip(db, user, zip_path) -> tuple:
    """Полный цикл для ZIP с инструкциями: распаковка → парсинг → сравнение →
    xlsx-отчёт в «Мои документы». Используется API-эндпоинтом и чат-веткой.
    Возвращает (MyDocuments, result)."""
    import tempfile
    import zipfile
    from datetime import datetime
    from pathlib import Path

    from config import settings
    from data.my_documents import MyDocuments
    from services.parsers import parse_file

    allowed = {".docx", ".doc", ".pdf", ".rtf", ".txt", ".odt"}
    zf = zipfile.ZipFile(zip_path)
    docs: list[tuple[str, str]] = []
    errors: list[str] = []
    with tempfile.TemporaryDirectory(prefix="ot_dedup_") as tmpdir:
        for info in zf.infolist():
            if info.is_dir() or len(docs) >= 500:
                continue
            fname = Path(info.filename).name
            suffix = Path(fname).suffix.lower()
            if not fname or suffix not in allowed:
                continue
            p = Path(tmpdir) / f"{len(docs)}{suffix}"
            p.write_bytes(zf.read(info))
            try:
                parsed = parse_file(p)
                if (parsed.text or "").strip():
                    docs.append((fname, parsed.text))
                else:
                    errors.append(fname)
            except Exception as e:
                logger.warning("[OT-DEDUP] {} не распарсился: {}", fname, e)
                errors.append(fname)
    if len(docs) < 2:
        raise ValueError("В архиве меньше двух читаемых инструкций (docx/doc/pdf/rtf/txt)")

    result = compare_documents(docs)
    result["unreadable"] = errors

    settings.docs_generated.mkdir(parents=True, exist_ok=True)
    out = settings.docs_generated / f"ot_dedup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out.write_bytes(build_dedup_xlsx(result))
    rec = MyDocuments(
        user_id=user.id,
        title=f"Дубликаты инструкций ОТ ({result['files']} файлов, {result['duplicates']} пар ≥80%)",
        template_key="ot_dedup",
        file_path=str(out),
        progress=100,
        status="ready",
        fields={"files": result["files"], "duplicates": result["duplicates"]},
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return rec, result
