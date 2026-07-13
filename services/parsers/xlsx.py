from __future__ import annotations

from pathlib import Path

from services.parsers.base import ParsedDocument, format_table


def parse_xlsx(path: str | Path) -> ParsedDocument:
    """Парсер Excel (.xlsx/.xlsm). Каждый лист → блок «Лист: <имя>», строки —
    через « | ». Подходит для штатных расписаний, графиков, реестров."""
    path = Path(path)
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise ValueError(
            "Для .xlsx нужен пакет openpyxl (pip install openpyxl)."
        ) from e

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for ws in wb.worksheets:
            if ws.max_row == 0:
                continue
            rows = [
                [("" if c is None else str(c)).strip() for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            table_lines = format_table(rows)
            if table_lines:
                parts.append(f"Лист: {ws.title}")
                parts.extend(table_lines)
    finally:
        wb.close()

    return ParsedDocument(
        text="\n\n".join(parts),
        title=path.stem,
        source_uri=str(path),
        source_type="local",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
