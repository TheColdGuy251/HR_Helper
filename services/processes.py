"""А10: приведение схем процессов «из разных программ Microsoft» к единому виду.

Экстрактор графа процесса из docx / pptx / xlsx (старые .doc/.ppt/.xls — через
LibreOffice-конвертацию): фигуры с текстом — узлы, коннекторы — рёбра
(направление по логической привязке stCxn/endCxn или по наконечникам стрелок),
короткие несоединённые надписи — роли блоков («УРП», «Ректор») либо метки
условий на стрелках («с командировкой», «отказ»). Затем послойный автолейаут
и рендер в SVG в едином стиле ТИУ.

Всё детерминированное, без LLM.
"""

from __future__ import annotations

import html
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path

from utils.logger import logger

# Триггер чат-команды: «приведи схему к единому виду», «перерисуй схему процесса»
PROCESS_REQUEST_RE = re.compile(
    r"(един\w+\s+(?:вид|стил)|схем\w+\s+процесс|перерису\w+\s+схем|стилизу\w+\s+схем)",
    re.IGNORECASE,
)

_EMU_MATCH = 700_000          # макс. расстояние конца стрелки до блока (~0.77 см)
_LABEL_MAX_CHARS = 60         # длиннее — точно блок, а не подпись
_XLSX_COL_EMU = 609_600       # ширина колонки Excel по умолчанию (64px)
_XLSX_ROW_EMU = 190_500       # высота строки по умолчанию (20px)


def _lc(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


@dataclass
class PNode:
    id: int
    text: str
    x: float
    y: float
    w: float
    h: float
    role: str | None = None
    dashed: bool = False

    @property
    def cx(self) -> float:
        return self.x + self.w / 2

    @property
    def cy(self) -> float:
        return self.y + self.h / 2


@dataclass
class PEdge:
    src: int
    dst: int
    label: str | None = None
    dashed: bool = False


@dataclass
class ProcessGraph:
    title: str | None
    nodes: list[PNode] = field(default_factory=list)
    edges: list[PEdge] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Сырые фигуры из XML
# ---------------------------------------------------------------------------


@dataclass
class _Shape:
    text: str
    x: float
    y: float
    w: float
    h: float
    is_conn: bool

    @property
    def cx(self) -> float:
        return self.x + self.w / 2

    @property
    def cy(self) -> float:
        return self.y + self.h / 2
    flip_h: bool = False
    flip_v: bool = False
    tail: bool = False       # наконечник на конце линии
    head: bool = False       # наконечник в начале линии
    dashed: bool = False
    st_id: str | None = None  # логическая привязка коннектора (pptx/xlsx)
    en_id: str | None = None
    shape_id: str | None = None


def _shape_text(el: ET.Element) -> str:
    """Абзацы текста фигуры (a:p → a:t), склеенные через перенос."""
    lines: list[str] = []
    for p in el.iter():
        if _lc(p.tag) != "p" or p.tag.startswith("{http://schemas.openxmlformats.org/wordprocessingml"):
            continue
        buf: list[str] = []
        for t in p.iter():
            if _lc(t.tag) == "t" and t.text:
                buf.append(t.text)
        line = _norm("".join(buf))
        if line:
            lines.append(line)
    # docx: текст лежит в w:p/w:t
    if not lines:
        for p in el.iter():
            if _lc(p.tag) != "p":
                continue
            buf = [t.text for t in p.iter() if _lc(t.tag) == "t" and t.text]
            line = _norm("".join(buf))
            if line and line not in lines:
                lines.append(line)
    # дедуп (mc:Fallback дублирует текст)
    seen: set[str] = set()
    uniq = [x for x in lines if not (x.lower() in seen or seen.add(x.lower()))]
    return "\n".join(uniq)


def _shape_flags(el: ET.Element) -> tuple[str, bool, bool, bool]:
    """(prst, tail, head, dashed) по prstGeom / tailEnd / headEnd / prstDash."""
    prst = ""
    tail = head = dashed = False
    for e in el.iter():
        lc = _lc(e.tag)
        if lc == "prstGeom" and not prst:
            prst = (e.get("prst") or "").lower()
        elif lc == "tailEnd" and (e.get("type") or "none") != "none":
            tail = True
        elif lc == "headEnd" and (e.get("type") or "none") != "none":
            head = True
        elif lc == "prstDash" and "dash" in (e.get("val") or ""):
            dashed = True
    return prst, tail, head, dashed


def _cxn_refs(el: ET.Element) -> tuple[str | None, str | None]:
    st = en = None
    for e in el.iter():
        lc = _lc(e.tag)
        if lc == "stCxn":
            st = e.get("id")
        elif lc == "endCxn":
            en = e.get("id")
    return st, en


def _sp_id(el: ET.Element) -> str | None:
    for e in el.iter():
        if _lc(e.tag) in ("cNvPr",):
            return e.get("id")
    return None


def _xfrm_geom(el: ET.Element) -> tuple[float | None, float | None, float, float, bool, bool]:
    """Геометрия из a:xfrm (pptx/фигуры): off + ext + flip."""
    for e in el.iter():
        if _lc(e.tag) == "xfrm":
            fx, fy = e.get("flipH") == "1", e.get("flipV") == "1"
            x = y = None
            w = h = 0.0
            for c in e:
                if _lc(c.tag) == "off":
                    x, y = float(c.get("x") or 0), float(c.get("y") or 0)
                elif _lc(c.tag) == "ext":
                    w, h = float(c.get("cx") or 0), float(c.get("cy") or 0)
            return x, y, w, h, fx, fy
    return None, None, 0.0, 0.0, False, False


def _shapes_from_pptx_like(root: ET.Element) -> list[_Shape]:
    """Фигуры слайда pptx (p:sp / p:cxnSp) — координаты в a:xfrm."""
    out: list[_Shape] = []
    for el in root.iter():
        lc = _lc(el.tag)
        if lc not in ("sp", "cxnSp"):
            continue
        x, y, w, h, fh, fv = _xfrm_geom(el)
        if x is None:
            continue
        prst, tail, head, dashed = _shape_flags(el)
        is_conn = lc == "cxnSp" or "connector" in prst or prst == "line" or "arrow" in prst
        st, en = _cxn_refs(el) if is_conn else (None, None)
        out.append(_Shape(
            text=_shape_text(el), x=x, y=y, w=w, h=h, is_conn=is_conn,
            flip_h=fh, flip_v=fv, tail=tail, head=head, dashed=dashed,
            st_id=st, en_id=en, shape_id=_sp_id(el),
        ))
    return out


def _shapes_from_docx(root: ET.Element) -> list[_Shape]:
    """Плавающие фигуры Word (wp:anchor / wp:inline) — абсолютный posOffset."""
    from services.parsers.docx import _anchor_geometry  # та же геометрия, что у парсера

    out: list[_Shape] = []
    for el in root.iter():
        if _lc(el.tag) not in ("anchor", "inline"):
            continue
        x, y, w, h, fh, fv = _anchor_geometry(el)
        if x is None or y is None:
            continue
        prst, tail, head, dashed = _shape_flags(el)
        is_conn = "connector" in prst or prst == "line" or "arrow" in prst
        out.append(_Shape(
            text=_shape_text(el), x=float(x), y=float(y), w=float(w), h=float(h),
            is_conn=is_conn, flip_h=fh, flip_v=fv, tail=tail, head=head, dashed=dashed,
        ))
    return out


def _xlsx_anchor_pt(el: ET.Element, colx, rowy) -> tuple[float, float] | None:
    col = row = coff = roff = None
    for c in el:
        lc = _lc(c.tag)
        if lc == "col":
            col = int(c.text or 0)
        elif lc == "colOff":
            coff = int(c.text or 0)
        elif lc == "row":
            row = int(c.text or 0)
        elif lc == "rowOff":
            roff = int(c.text or 0)
    if col is None or row is None:
        return None
    return colx(col) + (coff or 0), rowy(row) + (roff or 0)


def _shapes_from_xlsx_drawing(root: ET.Element, colx, rowy) -> list[_Shape]:
    """Стрелки/фигуры листа Excel (xdr:*CellAnchor). Координаты якорей — клетки,
    переводятся в EMU по РЕАЛЬНЫМ ширинам колонок и высотам строк (colx/rowy)."""
    out: list[_Shape] = []
    for anch in root.iter():
        if _lc(anch.tag) not in ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor"):
            continue
        frm = to = None
        body = None
        for c in anch:
            lc = _lc(c.tag)
            if lc == "from":
                frm = _xlsx_anchor_pt(c, colx, rowy)
            elif lc == "to":
                to = _xlsx_anchor_pt(c, colx, rowy)
            elif lc in ("sp", "cxnSp", "grpSp"):
                body = c
        if frm is None or body is None:
            continue
        if to is None:
            to = (frm[0] + 1_000_000, frm[1] + 400_000)
        x, y = min(frm[0], to[0]), min(frm[1], to[1])
        w, h = abs(to[0] - frm[0]), abs(to[1] - frm[1])
        prst, tail, head, dashed = _shape_flags(body)
        is_conn = _lc(body.tag) == "cxnSp" or "connector" in prst or prst == "line" or "arrow" in prst
        st, en = _cxn_refs(body) if is_conn else (None, None)
        # Направление xlsx-коннектора задаётся якорями from/to
        fh = frm[0] > to[0]
        fv = frm[1] > to[1]
        out.append(_Shape(
            text=_shape_text(body), x=x, y=y, w=w, h=h, is_conn=is_conn,
            flip_h=fh, flip_v=fv, tail=tail, head=head, dashed=dashed,
            st_id=st, en_id=en, shape_id=_sp_id(body),
        ))
    return out


def _xlsx_sheet_shape_sets(path: Path) -> list[list[_Shape]]:
    """Наборы фигур по листам Excel: БЛОКИ схемы — это заполненные ЯЧЕЙКИ
    (в т.ч. объединённые), стрелки — фигуры из drawingN.xml того же листа."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True)

    # sheet name → файл drawing (workbook.xml → sheet rels → drawing target)
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())

        def _read_xml(n: str) -> ET.Element | None:
            return ET.fromstring(z.read(n)) if n in names else None

        wb_root = _read_xml("xl/workbook.xml")
        wb_rels = _read_xml("xl/_rels/workbook.xml.rels")
        rid2target = {}
        if wb_rels is not None:
            for r in wb_rels:
                rid2target[r.get("Id")] = r.get("Target", "").lstrip("/")
        sheet_file: dict[str, str] = {}
        if wb_root is not None:
            for s in wb_root.iter():
                if _lc(s.tag) == "sheet":
                    rid = next((v for k, v in s.attrib.items() if k.endswith("}id")), None)
                    tgt = rid2target.get(rid, "")
                    if tgt and not tgt.startswith("xl/"):
                        tgt = "xl/" + tgt
                    sheet_file[s.get("name") or ""] = tgt
        drawing_of: dict[str, str] = {}
        for name, sf in sheet_file.items():
            if not sf or sf not in names:
                continue
            sroot = _read_xml(sf)
            rels = _read_xml(sf.rsplit("/", 1)[0] + "/_rels/" + sf.rsplit("/", 1)[1] + ".rels")
            drid = None
            for e in (sroot.iter() if sroot is not None else []):
                if _lc(e.tag) == "drawing":
                    drid = next((v for k, v in e.attrib.items() if k.endswith("}id")), None)
            if drid and rels is not None:
                for r in rels:
                    if r.get("Id") == drid:
                        t = r.get("Target", "")
                        drawing_of[name] = "xl/" + t.replace("../", "").lstrip("/")
        drawings_xml = {n: z.read(dr) for n, dr in drawing_of.items() if dr in names}

    sets: list[list[_Shape]] = []
    for ws in wb.worksheets:
        # Кумулятивные координаты колонок/строк в EMU (ширина в символах ≈ 7px/симв.)
        max_col = min(ws.max_column or 1, 60) + 4
        max_row = min(ws.max_row or 1, 300) + 4
        from openpyxl.utils import get_column_letter

        colw = []
        for i in range(1, max_col + 1):
            dim = ws.column_dimensions.get(get_column_letter(i))
            chars = dim.width if (dim and dim.width) else 8.43
            colw.append(int(round(chars * 7)) * 9525)
        rowh = []
        for i in range(1, max_row + 1):
            dim = ws.row_dimensions.get(i)
            pts = dim.height if (dim and dim.height) else 15.0
            rowh.append(int(round(pts * 12700)))
        cum_x = [0]
        for w in colw:
            cum_x.append(cum_x[-1] + w)
        cum_y = [0]
        for h in rowh:
            cum_y.append(cum_y[-1] + h)

        def colx(c: int) -> float:
            return cum_x[min(c, len(cum_x) - 1)]

        def rowy(r: int) -> float:
            return cum_y[min(r, len(cum_y) - 1)]

        shapes: list[_Shape] = []
        # Объединённые ячейки → блоки
        merged_cells: set[tuple[int, int]] = set()
        for rng in ws.merged_cells.ranges:
            v = ws.cell(rng.min_row, rng.min_col).value
            for rr in range(rng.min_row, rng.max_row + 1):
                for cc in range(rng.min_col, rng.max_col + 1):
                    merged_cells.add((rr, cc))
            text = _norm(str(v)) if v is not None else ""
            if not text:
                continue
            x, y = colx(rng.min_col - 1), rowy(rng.min_row - 1)
            shapes.append(_Shape(
                text=text, x=x, y=y,
                w=colx(rng.max_col) - x, h=rowy(rng.max_row) - y, is_conn=False,
            ))
        # Одиночные заполненные ячейки
        for row in ws.iter_rows(max_row=min(ws.max_row or 1, 300)):
            for cell in row:
                if cell.value is None or (cell.row, cell.column) in merged_cells:
                    continue
                text = _norm(str(cell.value))
                if not text:
                    continue
                x, y = colx(cell.column - 1), rowy(cell.row - 1)
                shapes.append(_Shape(
                    text=text, x=x, y=y,
                    w=colx(cell.column) - x, h=rowy(cell.row) - y, is_conn=False,
                ))
        # Стрелки листа
        xml = drawings_xml.get(ws.title)
        if xml:
            shapes.extend(_shapes_from_xlsx_drawing(ET.fromstring(xml), colx, rowy))
        if shapes:
            sets.append(shapes)
    return sets


def _collect_shape_sets(path: Path) -> list[list[_Shape]]:
    """Наборы фигур-кандидатов (по одному на «полотно»: документ/слайды/лист)."""
    suffix = path.suffix.lower()
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        if suffix == ".docx":
            root = ET.fromstring(z.read("word/document.xml"))
            return [_shapes_from_docx(root)]
        if suffix == ".pptx":
            return [
                _shapes_from_pptx_like(ET.fromstring(z.read(n)))
                for n in sorted(n for n in names if re.fullmatch(r"ppt/slides/slide\d+\.xml", n))
            ]
    if suffix in (".xlsx", ".xlsm"):
        return _xlsx_sheet_shape_sets(path)
    return []


# ---------------------------------------------------------------------------
# Сборка графа: блоки / роли / рёбра / метки условий
# ---------------------------------------------------------------------------


def _pt_rect_dist(px: float, py: float, s: _Shape) -> float:
    dx = max(s.x - px, 0, px - (s.x + s.w))
    dy = max(s.y - py, 0, py - (s.y + s.h))
    return (dx * dx + dy * dy) ** 0.5


def _looks_like_label(s: _Shape) -> bool:
    t = s.text.replace("\n", " ")
    return 0 < len(t) <= _LABEL_MAX_CHARS and len(t.split()) <= 7


def extract_process_graph(path: str | Path) -> ProcessGraph | None:
    """Извлекает граф процесса из файла со схемой (лучшее «полотно» файла).
    None — схема не распознана (нет блоков со стрелками, либо она — картинка)."""
    path = Path(path)
    if path.suffix.lower() in (".doc", ".ppt", ".xls"):
        from services.parsers.office_convert import convert_to_modern

        path = convert_to_modern(path)
    try:
        shape_sets = _collect_shape_sets(path)
    except Exception as e:
        logger.warning("[PROCESS] не удалось прочитать фигуры {}: {}", path.name, e)
        return None

    best: ProcessGraph | None = None
    for shapes in shape_sets:
        g = _assemble(shapes)
        if g and (best is None or len(g.nodes) > len(best.nodes)):
            best = g
    return best


def _assemble(shapes: list[_Shape]) -> ProcessGraph | None:
    # «? …» — комментарии-идеи на полях схем (не шаги процесса)
    texted = [
        s for s in shapes
        if s.text and not s.is_conn and not s.text.lstrip().startswith("?")
    ]
    conns = [s for s in shapes if s.is_conn]
    if len(texted) < 2 or not conns:
        return None

    # Надпись со строчной буквы («согласовано», «с командировкой») — метка
    # условия; короткая аббревиатура КАПСОМ («УРП», «ДЭФИ») — роль-дорожка.
    # Ни то ни другое не шаг процесса: исключаем из привязки концов стрелок.
    def _forced_label(s: _Shape) -> bool:
        t = s.text.replace("\n", " ")
        return (len(t) <= 30 and t[:1].islower()) or (len(t) <= 6 and t.isupper())

    id_by_spid = {s.shape_id: s for s in texted if s.shape_id}
    match_pool = [s for s in texted if not _forced_label(s)]

    # 1) Рёбра: логические привязки (pptx/xlsx) → геометрия концов линии.
    raw_edges: list[tuple[_Shape, _Shape, _Shape]] = []  # (src, dst, conn)

    def _nearest_block(pt: tuple[float, float]) -> tuple[_Shape | None, float]:
        """Ближайший блок к точке; короткие подписи (роли) — со штрафом, чтобы
        конец стрелки у заголовка колонки не «прилипал» к нему вместо шага."""
        best, bd = None, 1e18
        for b in match_pool:
            d = _pt_rect_dist(pt[0], pt[1], b)
            if _looks_like_label(b):
                d = d * 1.5 + 150_000
            if d < bd:
                best, bd = b, d
        return best, bd

    for c in conns:
        src = id_by_spid.get(c.st_id) if c.st_id else None
        dst = id_by_spid.get(c.en_id) if c.en_id else None
        if src is None or dst is None:
            p1 = (c.x + (c.w if c.flip_h else 0), c.y + (c.h if c.flip_v else 0))
            p2 = (c.x + (0 if c.flip_h else c.w), c.y + (0 if c.flip_v else c.h))
            # Наконечник tailEnd — на КОНЦЕ линии (p2): поток p1→p2; headEnd — наоборот.
            start_pt, end_pt = (p2, p1) if (c.head and not c.tail) else (p1, p2)
            s1, d1 = _nearest_block(start_pt)
            s2, d2 = _nearest_block(end_pt)
            if not s1 or not s2 or s1 is s2 or d1 > _EMU_MATCH or d2 > _EMU_MATCH:
                continue
            if not c.tail and not c.head and (s1.cy, s1.cx) > (s2.cy, s2.cx):
                s1, s2 = s2, s1  # линия без стрелки — сверху вниз / слева направо
            src, dst = s1, s2
        if src is dst:
            continue
        raw_edges.append((src, dst, c))

    if not raw_edges:
        return None

    connected = {id(s) for a, b, _ in raw_edges for s in (a, b)}

    # 2) Несоединённые короткие надписи: заголовок, метка ребра или роль блока.
    labels = [
        s for s in texted
        if _forced_label(s) or (id(s) not in connected and _looks_like_label(s))
    ]
    label_ids = {id(s) for s in labels}
    blocks = [s for s in texted if id(s) not in label_ids]
    if len(blocks) < 2:
        return None

    # Заголовок схемы: надпись ВЫШЕ всех блоков, ШИРЕ типового блока (роли —
    # узкие подписи над конкретным блоком) и начинающаяся в левой половине.
    title: str | None = None
    if labels:
        min_block_y = min(b.y for b in blocks)
        ws = sorted(b.w for b in blocks)
        median_w = ws[len(ws) // 2]
        mid_x = (min(b.x for b in blocks) + max(b.x + b.w for b in blocks)) / 2
        cands = [
            s for s in labels
            if s.y + s.h <= min_block_y and s.w >= 1.25 * median_w and s.x < mid_x
        ]
        if cands:
            top = min(cands, key=lambda s: s.y)
            title = top.text.replace("\n", " ")
            labels.remove(top)

    # Метка ребра — если её центр ближе к середине какой-то стрелки, чем к блокам.
    edge_label: dict[int, str] = {}
    roles: dict[int, str] = {}
    for lb in labels:
        lcx, lcy = lb.cx, lb.cy
        best_e, bde = None, 1e18
        for i, (_, _, c) in enumerate(raw_edges):
            mx, my = c.x + c.w / 2, c.y + c.h / 2
            d = ((mx - lcx) ** 2 + (my - lcy) ** 2) ** 0.5
            if d < bde:
                best_e, bde = i, d
        best_b, bdb = None, 1e18
        for b in blocks:
            d = _pt_rect_dist(lcx, lcy, b)
            if d < bdb:
                best_b, bdb = b, d
        # Метка со строчной буквы — условие перехода: ТОЛЬКО на стрелку
        # (в сетке Excel она всегда вплотную к какому-нибудь блоку, поэтому
        # сравнение расстояний с блоками тут не работает). Прочие подписи —
        # роль ближайшего блока.
        is_cond = lb.text[:1].islower()
        if is_cond:
            if best_e is not None and bde < _EMU_MATCH * 3:
                prev = edge_label.get(best_e)
                edge_label[best_e] = f"{prev} / {lb.text}" if prev else lb.text.replace("\n", " ")
            continue
        if best_e is not None and bde < bdb and bde < _EMU_MATCH * 2:
            prev = edge_label.get(best_e)
            edge_label[best_e] = f"{prev} / {lb.text}" if prev else lb.text.replace("\n", " ")
        elif best_b is not None and bdb < _EMU_MATCH * 3:
            # роль пишут НАД блоком или слева — принимаем ближайший блок
            if id(best_b) not in roles:
                roles[id(best_b)] = lb.text.replace("\n", " ")

    nodes: list[PNode] = []
    idx: dict[int, int] = {}
    for s in sorted(blocks, key=lambda b: (b.y, b.x)):
        idx[id(s)] = len(nodes)
        nodes.append(PNode(
            id=len(nodes), text=s.text, x=s.x, y=s.y, w=s.w, h=s.h,
            role=roles.get(id(s)), dashed=s.dashed,
        ))

    edges: list[PEdge] = []
    seen_e: set[tuple[int, int, str | None]] = set()
    for i, (a, b, c) in enumerate(raw_edges):
        if id(a) not in idx or id(b) not in idx:
            continue
        key = (idx[id(a)], idx[id(b)], edge_label.get(i))
        if key in seen_e:
            continue
        seen_e.add(key)
        edges.append(PEdge(src=idx[id(a)], dst=idx[id(b)], label=edge_label.get(i), dashed=c.dashed))

    return ProcessGraph(title=title, nodes=nodes, edges=edges)


# ---------------------------------------------------------------------------
# Автолейаут: слои по потоку (длиннейший путь), внутри слоя — по исходному y
# ---------------------------------------------------------------------------

_NODE_W = 230
_GAP_X = 110
_GAP_Y = 36
_BAND_GAP = 70       # зазор между лентами «змейки»
_MAX_COLS = 4        # колонок в ленте (дальше — перенос на следующую)
_FONT = 13
_LINE_H = 17
_PAD = 12
_CHARS_PER_LINE = 30


def _wrap(text: str, width: int = _CHARS_PER_LINE) -> list[str]:
    out: list[str] = []
    for raw_line in text.split("\n"):
        words = raw_line.split()
        cur = ""
        for w in words:
            cand = f"{cur} {w}".strip()
            if len(cand) > width and cur:
                out.append(cur)
                cur = w
            else:
                cur = cand
        if cur:
            out.append(cur)
    return out or [""]


def _layout(g: ProcessGraph) -> dict:
    n = len(g.nodes)
    adj: dict[int, list[int]] = {i: [] for i in range(n)}
    indeg = {i: 0 for i in range(n)}
    for e in g.edges:
        adj[e.src].append(e.dst)
        indeg[e.dst] += 1

    # Слой = длиннейший путь от источника (Kahn); циклы дожимаем по y.
    level = {i: 0 for i in range(n)}
    q = [i for i in range(n) if indeg[i] == 0] or [0]
    deg = dict(indeg)
    seen: set[int] = set()
    while q:
        v = q.pop(0)
        if v in seen:
            continue
        seen.add(v)
        for u in adj[v]:
            level[u] = max(level[u], level[v] + 1)
            deg[u] -= 1
            if deg[u] <= 0:
                q.append(u)
    for i in range(n):
        if i not in seen and any(i in adj[j] for j in seen):
            level[i] = max(level[i], 1)

    cols: dict[int, list[int]] = {}
    for i in range(n):
        cols.setdefault(level[i], []).append(i)
    for lv in cols:
        cols[lv].sort(key=lambda i: (g.nodes[i].y, g.nodes[i].x))

    pos: dict[int, tuple[float, float, float, float]] = {}  # id -> x,y,w,h (px)
    heights: dict[int, float] = {}
    for i, node in enumerate(g.nodes):
        lines = _wrap(node.text)
        h = len(lines) * _LINE_H + _PAD * 2 + (16 if node.role else 0)
        heights[i] = h

    # «Змейка»: длинная цепочка не должна растягиваться в километровую полосу —
    # уровни складываются в ленты по _MAX_COLS колонок, лента под лентой.
    n_levels = max(cols) + 1
    per_row = min(n_levels, _MAX_COLS)
    col_h = {lv: sum(heights[i] for i in ids) + _GAP_Y * (len(ids) - 1) for lv, ids in cols.items()}

    band_of = {lv: lv // _MAX_COLS for lv in cols}
    n_bands = max(band_of.values()) + 1
    band_h = {b: max(col_h[lv] for lv in cols if band_of[lv] == b) for b in range(n_bands)}
    band_y: dict[int, float] = {}
    y_cursor = 70.0
    for b in range(n_bands):
        band_y[b] = y_cursor
        y_cursor += band_h[b] + _BAND_GAP

    for lv, ids in sorted(cols.items()):
        b = band_of[lv]
        x = 40 + (lv % _MAX_COLS) * (_NODE_W + _GAP_X)
        y = band_y[b] + (band_h[b] - col_h[lv]) / 2
        for i in ids:
            pos[i] = (x, y, _NODE_W, heights[i])
            y += heights[i] + _GAP_Y

    width = 40 * 2 + per_row * _NODE_W + (per_row - 1) * _GAP_X
    height = y_cursor - _BAND_GAP + 50
    return {"pos": pos, "width": max(width, 460), "height": height}


# ---------------------------------------------------------------------------
# SVG в едином стиле ТИУ
# ---------------------------------------------------------------------------

_BLUE = "#1E40AF"
_INK = "#0F172A"
_MUTED = "#475569"


def render_process_svg(g: ProcessGraph) -> str:
    lt = _layout(g)
    pos = lt["pos"]
    out: list[str] = []
    out.append(
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {lt["width"]} {lt["height"]}" '
        f'width="{lt["width"]}" height="{lt["height"]}" '
        f'font-family="Segoe UI, Arial, sans-serif">'
    )
    out.append(
        '<defs><marker id="arr" viewBox="0 0 10 10" refX="9" refY="5" markerWidth="7" '
        'markerHeight="7" orient="auto-start-reverse">'
        f'<path d="M 0 0 L 10 5 L 0 10 z" fill="{_MUTED}"/></marker></defs>'
    )
    out.append(f'<rect x="0" y="0" width="{lt["width"]}" height="{lt["height"]}" fill="white"/>')
    if g.title:
        out.append(
            f'<text x="40" y="38" font-size="20" font-weight="700" fill="{_INK}">'
            f"{html.escape(g.title)}</text>"
        )
        out.append(f'<rect x="40" y="48" width="56" height="4" rx="2" fill="{_BLUE}"/>')

    # Рёбра — под блоками
    for e in g.edges:
        x1, y1, w1, h1 = pos[e.src]
        x2, y2, w2, h2 = pos[e.dst]
        if x2 > x1 + w1:  # вперёд по потоку: правая грань → левая грань
            sx, sy, ex, ey = x1 + w1, y1 + h1 / 2, x2, y2 + h2 / 2
            mx = (sx + ex) / 2
            d = f"M {sx:.0f} {sy:.0f} C {mx:.0f} {sy:.0f} {mx:.0f} {ey:.0f} {ex - 3:.0f} {ey:.0f}"
        elif y2 > y1 + h1 + 20:  # переход на ленту ниже («змейка»): вниз → вбок → вниз
            sx, sy, ex, ey = x1 + w1 / 2, y1 + h1, x2 + w2 / 2, y2
            my = (sy + ey) / 2
            d = f"M {sx:.0f} {sy:.0f} L {sx:.0f} {my:.0f} L {ex:.0f} {my:.0f} L {ex:.0f} {ey - 3:.0f}"
        elif x1 > x2 + w2:  # назад: дуга снизу
            sx, sy, ex, ey = x1, y1 + h1 / 2, x2 + w2, y2 + h2 / 2
            dip = max(y1 + h1, y2 + h2) + 40
            d = f"M {sx:.0f} {sy:.0f} C {sx - 60:.0f} {dip:.0f} {ex + 60:.0f} {dip:.0f} {ex + 3:.0f} {ey:.0f}"
        else:  # один слой: вертикально
            if y2 > y1:
                sx, sy, ex, ey = x1 + w1 / 2, y1 + h1, x2 + w2 / 2, y2
            else:
                sx, sy, ex, ey = x1 + w1 / 2, y1, x2 + w2 / 2, y2 + h2
            d = f"M {sx:.0f} {sy:.0f} L {ex:.0f} {ey - 3 if y2 > y1 else ey + 3:.0f}"
        dash = ' stroke-dasharray="6 4"' if e.dashed else ""
        out.append(
            f'<path d="{d}" fill="none" stroke="{_MUTED}" stroke-width="1.6"{dash} marker-end="url(#arr)"/>'
        )
        if e.label:
            lx, ly = (sx + ex) / 2, (sy + ey) / 2 - 6
            out.append(
                f'<text x="{lx:.0f}" y="{ly:.0f}" font-size="11" font-style="italic" '
                f'fill="{_BLUE}" text-anchor="middle">{html.escape(e.label)}</text>'
            )

    # Блоки
    for node in g.nodes:
        x, y, w, h = pos[node.id]
        dash = ' stroke-dasharray="6 4"' if node.dashed else ""
        out.append(
            f'<rect x="{x:.0f}" y="{y:.0f}" width="{w:.0f}" height="{h:.0f}" rx="10" '
            f'fill="white" stroke="{_BLUE}" stroke-width="1.6"{dash}/>'
        )
        ty = y + _PAD + _FONT
        if node.role:
            role = node.role.upper()
            if len(role) > 34:
                role = role[:33] + "…"
            out.append(
                f'<text x="{x + w / 2:.0f}" y="{y - 7:.0f}" font-size="11" font-weight="600" '
                f'fill="{_BLUE}" text-anchor="middle">{html.escape(role)}</text>'
            )
        for line in _wrap(node.text):
            out.append(
                f'<text x="{x + w / 2:.0f}" y="{ty:.0f}" font-size="{_FONT}" fill="{_INK}" '
                f'text-anchor="middle">{html.escape(line)}</text>'
            )
            ty += _LINE_H
    out.append("</svg>")
    return "".join(out)
